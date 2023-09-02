import os
import re
import json
import datetime
import shutil
from openpyxl import load_workbook

from modle3_text.server.config import seting
from modle3_text.server.unils import req


class Handel:
    """
    实现功能：
    1. 用户登录
    2. 用户注册
    3. 用户查看网盘内容 [登录后操作]
    4. 用户上传文件 [登录后操作]
    5. 用户下载文件 [登录后操作]
    6. 用户删除文件 [登录后操作]
    ...
    """

    def __init__(self, conn):
        self.conn = conn
        self.username = None

    @property
    def user_home_page(self):
        """用户文件路径"""
        return os.path.join(seting.USER_FOLDER_PATH, self.username)

    def send_data(self, **kwargs):
        req.send_data(self.conn, json.dumps(kwargs))

    def recv_save_file(self, file_path):
        req.recv_save_file(self.conn, file_path)

    def login(self, username, password):
        """用户登录"""
        wb = load_workbook(seting.DB_BASE_PATH)
        sheet = wb.worksheets[0]
        success = False
        for row in sheet.rows:
            if row[0].value == username and row[1].value == password:
                success = True
                break
        if success:
            self.send_data(status=True, data="登录成功")
            self.username = username

        else:
            self.send_data(status=False, error="用户名或密码错误")

    def register(self, username, password):
        """用户注册"""
        wb = load_workbook(seting.DB_BASE_PATH)
        sheet = wb.worksheets[0]

        exists = False
        for row in sheet.rows:
            if row[0].value == username:
                exists = True
                break
        if exists:
            # 给客户端回复用户名已存在
            self.send_data(status=False, data="用户名已存在")
            return

        # 给客户端回复注册成功
        max_row = sheet.max_row
        data_list = [username, password, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        for i, item in enumerate(data_list, 1):
            cell = sheet.cell(max_row + 1, i)
            cell.value = item
        wb.save(seting.DB_BASE_PATH)

        # 创建用户目录
        user_folder = os.path.join(seting.USER_FOLDER_PATH, username)
        os.makedirs(user_folder)
        self.send_data(status=True, data="注册成功")
        self.login(username, password)

    def show(self, user_file_path=None):
        if not self.username:
            self.send_data(status=False, error="未登录系统，请登录后查看")
            return

        if not user_file_path:
            data_path = '\n'.join(os.listdir(self.user_home_page))
            if data_path:
                self.send_data(status=True, data=data_path)
                return
            else:
                self.send_data(status=True, data="用户目录下暂时没有文件")
                return

        targe_user_file_path = os.path.join(self.user_home_page, user_file_path)
        if not os.path.exists(targe_user_file_path):
            self.send_data(status=False, error="路径不存在")
            return
        if not os.path.isdir(targe_user_file_path):
            self.send_data(status=False, error="文件夹不存在")
            return
        data_path_inner = '\n'.join(os.listdir(targe_user_file_path))
        if data_path_inner:
            self.send_data(status=True, data=data_path_inner)
            return
        else:
            self.send_data(status=True, data=f"`{user_file_path}`目录下暂时没有文件")
            return

    def upload(self, file_path):
        if not self.username:
            self.send_data(status=False, error="未登录系统，请登录后查看")
        targe_file_path = os.path.join(self.user_home_page, file_path)
        folder = os.path.dirname(targe_file_path)
        if not os.path.exists(folder):
            os.mkdir(folder)
        self.send_data(status=True, data='开始上传')
        self.recv_save_file(targe_file_path)

    def download(self, file_path, seek=0):
        if not self.username:
            self.send_data(status=False, error="未登录系统，请登录后查看")
        targe_file_path = os.path.join(self.user_home_page, file_path)
        if not os.path.exists(targe_file_path):
            self.send_data(status=False, data=f'文件{targe_file_path}不存在')
            return
        self.send_data(status=True, data='开始下载')

        seek = int(seek)
        targe_size = os.stat(targe_file_path).st_size
        req.send_file(self.conn, targe_size - seek, targe_file_path, seek)

    def delete(self, file_path):
        if not self.username:
            self.send_data(status=False, errno="未登录系统，请登陆后查看")
        target_file_path = os.path.join(self.user_home_page, file_path)

        if os.path.isdir(target_file_path):
            shutil.rmtree(target_file_path)
            self.send_data(status=True, data='已删除目标文件夹')
        else:
            os.remove(target_file_path)
            self.send_data(status=True, data='已删除目标文件')

    def mkdir(self, file_path):
        if not self.username:
            self.send_data(status=False, errno="未登录系统，请登陆后查看")
        targe_file_path = os.path.join(self.user_home_page, file_path)
        if os.path.exists(targe_file_path) and os.path.isdir(targe_file_path):
            self.send_data(status=False, data='文件夹已存在')
            return
        else:
            os.mkdir(targe_file_path)
            self.send_data(status=True, data=f'已创建文件夹{file_path}')

    def execute(self):
        """
        每次客户端发来请求，触发此方法。
        :return: False，关闭连接；True，继续处理请求
        """
        data = req.recv_data(self.conn).decode('utf-8')
        # print(data)
        if data.upper() == 'Q':
            print('客户端已被释放')
            return False

        print('接收到客户端请求：{}'.format(data))

        function_map = {
            'login': self.login,
            'register': self.register,
            'ls': self.show,
            'upload': self.upload,
            'download': self.download,
            'delete': self.delete,
            'mkdir': self.mkdir,
        }

        # 切割空格成列表 '命令 用户名 密码' -> [命令,用户名，密码]
        data, *args = re.split(r'\s+', data)
        data_method = function_map[data]
        data_method(*args)
        return True
